from openpyxl import load_workbook
from copy import copy
import numpy as np


# Funkcja do załadowania danych z pliku bazy danych i wrzucenia ich do zmiennej
def load_data(file_path: str):
    workbook = load_workbook(file_path)
    sheet = workbook.sheetnames[0]
    data = []
    criteria = []
    for col in workbook[sheet].iter_cols(values_only=True):
        data.append(list(col[1:]))
        criteria.append(col[0])
    return criteria, data


# Funkcja do sortowania danych (w naszym przypadku po cenie mieszkania)
def sorting_data(data: list):
    data_sorted = [list(x) for x in zip(*sorted(zip(*data), key=lambda x: x[0], reverse=True))]
    return data_sorted


# Funkcja do pobrania ilości przedziałów dla każdego kryterium od użytkownika
def collect_no_of_sections(criteria: list):
    no_of_sections = []
    for name in criteria[:-2]:
        input_string = "Podaj ilość przedziałów dla kryterium '" + str(name) + "': "
        no_of_sections.append(int(input(input_string)))
    no_of_sections.append(2)
    no_of_sections.append(2)

    if not all(val >= 2 for val in no_of_sections):
        print("Każde kryterium musi mieć więcej niż jeden przedział")
        no_of_sections = collect_no_of_sections(criteria)
    return no_of_sections


# Funkcja do dzielenia każdego kryterium na równe przedziały
def create_sections(data:list, criteria:list, no_of_sections:list):
    min_values = [min(value_list) for value_list in data]
    max_values = [max(value_list) for value_list in data]

    increment = [(max_values[i] - min_values[i]) / no_of_sections[i] for i in range(len(data))]

    sections = [[val] for val in min_values]
    for i in range(len(data)):
        incrementing_number = min_values[i]
        for _ in range(no_of_sections[i]-1):
            incrementing_number += increment[i]
            sections[i].append(incrementing_number)
        sections[i].append(max_values[i])

    for i in range(3, len(data)):
        sections[i] = sections[i][::-1]

    sections_as_tuples = []
    for sublist in sections:
        tuple_list = [(sublist[i], sublist[i+1]) for i in range(len(sublist)-1)]
        sections_as_tuples.append(tuple_list)

    return sections, sections_as_tuples


# Funkcja do wpisywania wartości funkcji użyteczności
def set_values_of_usability_functions(sections: list, criteria: list):
    values_of_usability_functions = []
    for i in range(len(sections)):
        value_list = []
        print("Dla kryterium '", criteria[i], "' wybierz wartości funkcji użyteczności dla punktów:", sep='')
        for j in range(len(sections[i]) - 1):
            point_string = str(criteria[i]) + "; " + str(sections[i][j]) + ": "
            value_list.append(float(input(point_string)))
        value_list.append(0)
        values_of_usability_functions.append(value_list)

    sum_for_ideal_point = sum([val[0] for val in values_of_usability_functions])
    if sum_for_ideal_point != 1:
        for i in range(len(values_of_usability_functions)):
            for j in range(len(values_of_usability_functions[i])):
                values_of_usability_functions[i][j] /= sum_for_ideal_point

    return values_of_usability_functions


# Funkcja wyznaczająca funkcje użyteczności
def usability_functions(sections: list, usability_values: list):
    a_func = []
    b_func = []
    for i in range(len(usability_values)):
        each_a = []
        each_b = []
        for j in range(len(usability_values[i]) - 1):
            a = (usability_values[i][j+1] - usability_values[i][j]) / (sections[i][j+1] - sections[i][j])
            each_a.append(a)
            each_b.append(usability_values[i][j+1] - a * sections[i][j+1])
        a_func.append(each_a)
        b_func.append(each_b)

    return a_func, b_func


# Funkcja wyznaczająca wartość ux
def function_value(flat_idx: int, a: list, b: list, data: list, sections_tuple: list):
    score = 0
    for i in range(len(data)):
        for section in sections_tuple[i]:
            if min(section) <= data[i][flat_idx] <= max(section):
                section_idx = sections_tuple[i].index(section)
                score += a[i][section_idx] * data[i][flat_idx] + b[i][section_idx]
                break
    return score


# Funkcja wyznaczająca scoring wszystkich mieszkań
def calculate_scoring(a: list, b: list, data: list, sections_tuple: list):
    scoring = []
    for idx in range(len(data[0])):
        scoring.append(function_value(idx, a, b, data, sections_tuple))
    return scoring


# Funkcja wyznaczająca ranking mieszkań
def calculate_ranking(scoring: list, data: list):
    data_copy = copy(data)
    data_copy.append(scoring)
    data_ranked = [list(x) for x in zip(*sorted(zip(*data_copy), key=lambda x: x[-1], reverse=True))]
    return data_ranked


# Funkcja wykonująca cały algorytm UTA STAR
def uta(file_path: str, no_of_sections=None, usability_values=None):
    criteria, data = load_data(file_path)
    sorted_data = sorting_data(data)
    if no_of_sections is None:
        no_of_sections = collect_no_of_sections(criteria)
    sections, sections_tuples = create_sections(sorted_data, criteria, no_of_sections)
    if usability_values is None:
        usability_values = set_values_of_usability_functions(sections, criteria)
    sum_for_ideal_point = sum([val[0] for val in usability_values])
    if sum_for_ideal_point != 1:
        for i in range(len(usability_values)):
            for j in range(len(usability_values[i])):
                usability_values[i][j] /= sum_for_ideal_point
    a, b = usability_functions(sections, usability_values)
    scoring = calculate_scoring(a, b, sorted_data, sections_tuples)
    ranking = calculate_ranking(scoring, sorted_data)

    rank = np.array([])
    value = np.array([])

    for i in range(len(data[0])):
        criteria_idx = 0
        print_string = "Miejsce " + str(i+1) + " (wartość funkcji użyteczności: " + str(ranking[-1][i]) + \
                       ") zajmuje mieszkanie z parametrami: "
        value = np.append(value, ranking[-1][i])
        tmp_lst = np.array([[]])

        for lst in ranking[:-1]:
            print_string += str(criteria[criteria_idx]) + ": " + str(lst[i]) + ", "
            criteria_idx += 1
            tmp_lst = np.append(tmp_lst, lst[i])
        # print(print_string)
        rank = np.append(rank, tmp_lst, axis=0)

    value = np.vstack(value)
    rank = rank.reshape((len(value), 7))
    rank = np.append(rank, value, axis=1)

    return rank


if __name__ == '__main__':
    przedzialy = None
    wartosci = None

    # jeśli chcesz żeby program nie pytał cię o ilość przedziałów i wartości funkcji użyteczności to zakomentuj następne
    # 2 linijki, w przeciwnym wypadku możesz wpisywać tutaj wartości manualnie
    #przedzialy = [5,2,3,3,2,2,2] # każdy element chyba powinien być większy lub równy 2 ale nie jestem pewny
    #wartosci = [[3, 0.3, 0.1, 0.07, 0.05, 0],[0.2, 0.12, 0],[0.2, 0.1, 0.07, 0],[3, 0.2, 0.1, 0],[0.1, 0.07, 0],[0.1, 0.03, 0],[0.1, 0.09, 0]]
    przedzialy = [2,2,2,2,2,2,2]
    wartosci = [[5,0,0],[0,0,0],[0,0,0],[5,0,0],[0,0,0],[0,0,0],[0,0,0]]
    uta("SWD_baza_danych.xlsx", przedzialy, wartosci)
